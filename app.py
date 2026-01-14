import streamlit as st
import google.generativeai as genai
import tempfile
import os
import pandas as pd
import json
import io
import openpyxl
from datetime import datetime
from fpdf import FPDF
import re
import traceback
import time
from google.generativeai.types import GenerationConfig
import zipfile

# --- CONSTANTS & CONFIGURATION ---
COMPANY_NAME = "Rüttenscheid Baukonzepte GmbH"

# Helper function for German number formatting
def format_german_number(value, decimals=2):
    """Format number in German style: 1.234.567,89"""
    if pd.isna(value) or value is None:
        return "0,00"
    try:
        value = float(value)
        # Python's default format uses comma for thousands and dot for decimal (English format)
        # Example: 1234567.89 -> "1,234,567.89"
        formatted = f"{value:,.{decimals}f}"

        # Now swap to German format:
        # Step 1: Replace comma (thousand separator) with dot
        formatted = formatted.replace(',', '.')
 
        parts = formatted.rsplit('.', 1)  # Split from right, only once
        if len(parts) == 2:
            # parts[0] = "1.234.567", parts[1] = "89"
            formatted = parts[0] + ',' + parts[1]

        return formatted
    except:
        return "0,00"

# Master AI Prompt - Handles ALL file types
MASTER_EXTRACTION_PROMPT = """
Du bist ein erfahrener Baukalkulator mit 25 Jahren Erfahrung bei der Rüttenscheid Baukonzepte GmbH.

DEINE AUFGABE:
Analysiere das hochgeladene Dokument (GAEB-Datei, PDF, Excel, Word) und extrahiere ALLE Bauleistungspositionen mit vollständiger Kalkulation.

⚠️ WICHTIG - ZEICHENCODIERUNG:
Falls du korrupte Zeichen wie �, �, oder ähnliche Symbole siehst, korrigiere sie zu korrekten deutschen Umlauten:

B�den → Böden
M�rtel → Mörtel
F�hren → Führen
Oberfl�che → Oberfläche
Verg�tung → Vergütung
Schl�ssel → Schlüssel
Nutze den Kontext, um die korrekte Bedeutung zu erkennen und deutsche Umlaute (ä, ö, ü, ß) richtig einzusetzen.
WICHTIGE ANWEISUNGEN:

POSITIONSIDENTIFIKATION:
✓ POSITION IST: Jede Zeile mit einer Leistungsbeschreibung UND Mengenangabe
✓ Beispiele: "Einrichten Baustelle", "Abbruch Mauerwerk", "Lieferung Beton C25/30"
✗ KEINE POSITION: Reine Überschriften, Abschnittstitel, Bemerkungen ohne Menge
✗ Beispiele: "Kogr. 391", "Baustelleneinrichtung" (nur Titel), "Hinweis: ..."

DATEN EXTRAKTION:

Positionsnummer: Extrahiere die EXAKTE Nummer (z.B. "01.01.0010", "0010", "1.2.3")
Falls keine Nummer vorhanden: Vergebe fortlaufende Nummern (0001, 0002, 0003, ...)

Beschreibung: Vollständige Leistungsbeschreibung (nicht nur Titel!)
Beispiel: "Erdarbeiten, Oberboden abtragen und zwischenlagern, Schichtdicke bis 30 cm"

Menge: Die EXAKTE numerische Menge (z.B. 150.5, 25, 1)
Falls nicht angegeben: 1.0

Einheit: Die EXAKTE Maßeinheit (m, m², m³, St, Std, t, kg, Psch, to, l, etc.)
Falls nicht angegeben: "Psch" (Pauschale)

PREISKALKULATION (SEHR WICHTIG):
Berechne für JEDE Position einen realistischen, marktgerechten Einheitspreis (netto, ohne MwSt.).

Berücksichtige:

Materialkosten (Einkauf, Beschaffung, Verschnitt 3-5%)
Lohnkosten (Facharbeiter ~45-55 EUR/Std, Hilfskraft ~35-45 EUR/Std)
Gerätekosten (Abschreibung, Betriebsstoffe, Vorhalten)
Transportkosten (An- und Abfahrt, Materialtransport)
Baustellengemeinkosten (BGK: ~8-12%)
Allgemeine Geschäftskosten (AGK: ~5-8%)
Wagnis & Gewinn (W&G: ~3-5%)
PREISBEISPIELE zur Orientierung:

Baustelleneinrichtung: 1.500,00 - 5.000,00 EUR (Pauschale)
Erdaushub Bagger: 8,50-15,75 EUR/m³
Beton C25/30 liefern und einbauen: 180,25-250,80 EUR/m³
Mauerwerk Ziegel herstellen: 85,50-120,75 EUR/m²
Bewehrung liefern und verlegen: 1.200,00-1.800,00 EUR/t
Schalung herstellen und abbrechen: 35,25-65,80 EUR/m²
Dachziegel decken: 45,60-75,90 EUR/m²
Putz innen auftragen: 25,50-40,75 EUR/m²
Fliesen verlegen: 40,80-80,50 EUR/m²

⚠️ KRITISCH WICHTIG:
• JEDE Position MUSS einen Preis > 0 haben!
• Keine Position darf unit_price: 0.0 haben!
• Preise MÜSSEN Cent-Beträge enthalten (z.B. 45.50, 125.75, 1250.25)
• NIEMALS runde Zahlen wie 45.00, 100.00, 250.00 verwenden!
• Realistische Preise haben IMMER Nachkommastellen!

AUSGABEFORMAT:
Gib NUR ein valides JSON-Array zurück.
KEINE Markdown-Formatierung, KEINE Erklärungen, KEINE zusätzlichen Texte.
NUR das reine JSON-Array.

Format:
[
{"pos": "0010", "description": "Vollständige Beschreibung", "quantity": 10.5, "unit": "m²", "unit_price": 45.50},
{"pos": "0020", "description": "Nächste Position...", "quantity": 25.0, "unit": "m³", "unit_price": 125.75},
{"pos": "0030", "description": "Weitere Position...", "quantity": 15.0, "unit": "t", "unit_price": 1450.25}
]

QUALITÄTSKRITERIEN:
✓ Mindestens 5 Zeichen pro Beschreibung
✓ Alle Mengen müssen Zahlen sein (keine Texte)
✓ Alle Preise müssen > 0 sein
✓ Einheiten müssen sinnvolle Baueinheiten sein

Starte jetzt die Analyse und gib NUR das JSON-Array zurück!
"""

# Secondary prompt if first attempt returns zero prices
PRICING_CORRECTION_PROMPT = """
Du bist Kalkulations-Experte. Die folgenden Positionen haben teilweise Preis 0.00 oder unrealistische Preise.

AUFGABE: Korrigiere ALLE Preise und stelle sicher, dass jede Position einen realistischen Marktpreis hat.

REGELN:

JEDE Position MUSS unit_price > 0 haben
Preise müssen marktgerecht und wettbewerbsfähig sein
Berücksichtige Material, Lohn, Geräte, Transport, BGK, AGK, W&G
Orientierung an typischen Baupreisen 2024/2025
PREISREFERENZEN:

Einfache Arbeiten (Aushub, Abbruch): 10,50-50,75 EUR
Mittlere Komplexität (Mauerwerk, Schalung): 50,25-150,80 EUR
Hohe Komplexität (Beton, Bewehrung): 150,50-300,75 EUR
Spezialleistungen (Abdichtung, Dämmung): 300,00+ EUR

⚠️ WICHTIG:
• Alle Preise MÜSSEN Cent-Beträge haben (z.B. 45.50, 125.75, 250.25)
• NIEMALS nur runde Zahlen wie 100.00 oder 250.00!
• Realistische Baupreise haben IMMER Nachkommastellen!
Eingabe-Positionen:
{positions_json}

AUSGABE: NUR das vollständige JSON-Array mit ALLEN Positionen und korrigierten Preisen.
Keine Erklärungen, nur JSON!
"""

# --- AI EXTRACTION FUNCTIONS ---
def get_mime_type(file_path):
    """
    Determine MIME type based on file extension.
    """
    ext = os.path.splitext(file_path)[1].lower()
    
    mime_types = {
        # Documents
        '.pdf': 'application/pdf',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.doc': 'application/msword',
        '.txt': 'text/plain',
        
        # Spreadsheets
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.xls': 'application/vnd.ms-excel',
        
        # GAEB files (German construction standard) - all as text/plain for Gemini compatibility
        '.d81': 'text/plain', '.d82': 'text/plain', '.d83': 'text/plain',
        '.d84': 'text/plain', '.d85': 'text/plain', '.d86': 'text/plain', '.d90': 'text/plain',
        '.x81': 'text/plain', '.x82': 'text/plain', '.x83': 'text/plain',
        '.x84': 'text/plain', '.x85': 'text/plain', '.x86': 'text/plain', '.x90': 'text/plain',
        '.p81': 'text/plain', '.p82': 'text/plain', '.p83': 'text/plain',
        '.p84': 'text/plain', '.p85': 'text/plain', '.p86': 'text/plain', '.p90': 'text/plain',
    }
    
    return mime_types.get(ext, 'application/octet-stream')

def call_ai_with_retry(model, contents, max_retries=3, initial_delay=5):
    """
    Call AI API with exponential backoff retry logic and automatic model switching.
    Tries alternative models when encountering 503 (overloaded) or 429 (quota exceeded) errors.
    Returns tuple: (response, model_used)
    """
    # Define available models in priority order (strongest first, then faster fallbacks)
    available_models = [
        'gemini-2.5-flash',       # Best balance: fast + capable
        'gemini-2.5-flash-lite',  # Fast and capable
        'gemini-2.5-pro',         # Most capable 2.5
        'gemini-2.0-flash',       # Reliable fallback
        'gemini-2.0-flash-lite',  # Fast fallback
        'gemini-3-pro',           # Newest pro (if available)
        'gemini-3-flash',         # Newest flash (if available)
    ]
    
    # Start with the requested model, then try others if needed
    if model in available_models:
        # Move requested model to front
        models_to_try = [model] + [m for m in available_models if m != model]
    else:
        models_to_try = available_models
    
    last_error = None
    
    for model_idx, current_model in enumerate(models_to_try):
        print(f"🤖 Trying model: {current_model}")
        
        for attempt in range(max_retries):
            try:
                model = genai.GenerativeModel(current_model)
                response = model.generate_content(contents)
                if model_idx > 0:
                    print(f"✅ Successfully switched to model: {current_model}")
                return response, current_model
                
            except Exception as e:
                last_error = e
                error_str = str(e)
                
                # Check for 503/overloaded errors
                if '503' in error_str or 'overloaded' in error_str.lower():
                    print(f"⚠️ Model {current_model} is overloaded (503)")
                    if model_idx < len(models_to_try) - 1:
                        print(f"🔄 Switching to next model...")
                        break
                    elif attempt < max_retries - 1:
                        wait_time = initial_delay * (2 ** attempt)
                        print(f"⚠️ All models tried. Retrying in {wait_time}s... (Attempt {attempt + 1}/{max_retries})")
                        time.sleep(wait_time)
                    else:
                        continue
                
                # Check for 429/quota/rate limit errors
                elif '429' in error_str or 'quota' in error_str.lower() or 'RESOURCE_EXHAUSTED' in error_str:
                    # Check if this is a token quota error (per-minute limit)
                    if 'input_token_count' in error_str or 'GenerateContentInputTokensPerModelPerMinute' in error_str:
                        print(f"⚠️ Token quota exceeded for {current_model}")
                        
                        # Extract retry delay from error message
                        retry_delay = 60  # Default
                        match = re.search(r'retry in (\d+(?:\.\d+)?)', error_str, re.IGNORECASE)
                        if match:
                            retry_delay = float(match.group(1))
                        
                        if model_idx < len(models_to_try) - 1:
                            print(f"🔄 Switching to next model...")
                            break
                        elif attempt < max_retries - 1:
                            print(f"⚠️ Waiting {retry_delay:.0f}s before retry...")
                            time.sleep(retry_delay)
                        else:
                            continue
                    else:
                        # Daily request quota exceeded - try next model
                        print(f"⚠️ Request quota exceeded for {current_model}")
                        if model_idx < len(models_to_try) - 1:
                            print(f"🔄 Switching to next model...")
                            break
                        else:
                            continue
                
                # Check for 500/502/504 server errors
                elif '500' in error_str or '502' in error_str or '504' in error_str:
                    if attempt < max_retries - 1:
                        wait_time = initial_delay * (2 ** attempt)
                        print(f"⚠️ API error (5xx). Retrying in {wait_time}s... (Attempt {attempt + 1}/{max_retries})")
                        time.sleep(wait_time)
                    else:
                        if model_idx < len(models_to_try) - 1:
                            print(f"🔄 Trying next model...")
                            break
                        else:
                            raise
                
                # Other errors - try next model or raise
                else:
                    if model_idx < len(models_to_try) - 1:
                        print(f"⚠️ Error with {current_model}: {error_str[:100]}")
                        print(f"🔄 Trying next model...")
                        break
                    else:
                        raise
    
    # If we got here, all models failed
    if last_error:
        raise last_error
    else:
        raise Exception("All models exhausted without successful response")

def safe_remove_file(file_path, max_retries=3, delay=0.5):
    """
    Safely remove a file with retry logic for Windows file locking issues.
    """
    for attempt in range(max_retries):
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
            return True
        except PermissionError:
            if attempt < max_retries - 1:
                time.sleep(delay)
            else:
                print(f"⚠️ Could not delete temporary file: {file_path}")
                return False
        except Exception as e:
            print(f"⚠️ Error deleting file: {e}")
            return False
    return False

def sanitize_filename(filename):
    """
    Sanitize filename by replacing special characters and umlauts.
    """
    # Replace German umlauts
    replacements = {
        'ä': 'ae', 'ö': 'oe', 'ü': 'ue', 'ß': 'ss',
        'Ä': 'Ae', 'Ö': 'Oe', 'Ü': 'Ue',
        ' ': '_',  # Replace spaces with underscores
        '/': '-', '\\': '-', ':': '-', '*': '-', '?': '-',
        '"': '-', '<': '-', '>': '-', '|': '-'
    }
    
    sanitized = filename
    for char, replacement in replacements.items():
        sanitized = sanitized.replace(char, replacement)
    
    # Remove any remaining non-ASCII characters
    sanitized = ''.join(c for c in sanitized if ord(c) < 128)
    
    return sanitized

def check_excel_structure(file_path):
    """
    Check if Excel file has the expected structure for direct extraction.
    Returns True if structure matches (Position in col A OR Ordnungszahl in col B).
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Strategy 1: Check for "Position" in column A
        position_count = 0
        for row in range(1, min(50, sheet.max_row + 1)):
            cell_value = sheet[f'A{row}'].value
            if cell_value:
                cell_str = str(cell_value).strip().lower()
                if "position" in cell_str or cell_str.startswith("pos"):
                    position_count += 1
                    print(f"[OK] Found position marker at row {row}: '{cell_value}'")
                    if position_count >= 1:
                        workbook.close()
                        print(f"[OK] Found {position_count} Position(s) in column A - structure detected")
                        return True

        # Strategy 2: Check for "Ordnungszahl" header structure
        for row in range(1, min(20, sheet.max_row + 1)):
            cell_b = str(sheet[f'B{row}'].value or "").strip().lower()
            cell_c = str(sheet[f'C{row}'].value or "").strip().lower()
            cell_d = str(sheet[f'D{row}'].value or "").strip().lower()

            # Look for the typical German structure: Ordnungszahl | Kurztext | Langtext
            if "ordnungszahl" in cell_b and "kurztext" in cell_c and "langtext" in cell_d:
                workbook.close()
                print(f"[OK] Found German LV structure at row {row}: Ordnungszahl | Kurztext | Langtext")
                return True

        workbook.close()
        print(f"⚠️ No matching structure found")
        return False
    except Exception as e:
        print(f"Error checking Excel structure: {e}")
        return False

def extract_positions_from_structured_excel(file_path):
    """
    Extract positions from Excel file with known structure.
    Supports two formats:
    1. Position in column A (old format)
    2. Typ | Ordnungszahl | Kurztext | Langtext (German LV format)
    Returns list of dictionaries with position data.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        positions = []

        # Find header row for German LV format
        header_row = None
        for row in range(1, min(20, sheet.max_row + 1)):
            cell_b = str(sheet[f'B{row}'].value or "").strip().lower()
            cell_c = str(sheet[f'C{row}'].value or "").strip().lower()
            cell_d = str(sheet[f'D{row}'].value or "").strip().lower()

            if "ordnungszahl" in cell_b and "kurztext" in cell_c and "langtext" in cell_d:
                header_row = row
                print(f"[OK] Found header row at row {row}")
                break

        # Extract using German LV format
        if header_row:
            for row in range(header_row + 1, sheet.max_row + 1):
                typ = str(sheet[f'A{row}'].value or "").strip().lower()

                # Only extract rows marked as "Position"
                if typ == "position":
                    ordnungszahl = sheet[f'B{row}'].value
                    kurztext = sheet[f'C{row}'].value
                    langtext = sheet[f'D{row}'].value
                    menge = sheet[f'E{row}'].value
                    einheit = sheet[f'F{row}'].value

                    # Skip if no description
                    if not kurztext and not langtext:
                        continue

                    position = {
                        "ordnungszahl": str(ordnungszahl) if ordnungszahl else "",
                        "kurztext": str(kurztext) if kurztext else "",
                        "langtext": str(langtext) if langtext else "",
                        "menge": menge if menge else 1.0,
                        "einheit": str(einheit) if einheit else "Psch"
                    }

                    positions.append(position)
                    print(f"[OK] Row {row}: {position['kurztext'][:50]}")

        # Fallback: old format with "Position" in column A
        else:
            for row in range(1, sheet.max_row + 1):
                cell_value = sheet[f'A{row}'].value

                if cell_value:
                    cell_str = str(cell_value).strip().lower()
                    if "position" in cell_str or cell_str.startswith("pos"):
                        ordnungszahl = sheet[f'B{row}'].value
                        kurztext = sheet[f'C{row}'].value
                        langtext = sheet[f'D{row}'].value
                        menge = sheet[f'E{row}'].value
                        einheit = sheet[f'F{row}'].value

                        position = {
                            "ordnungszahl": str(ordnungszahl) if ordnungszahl else "",
                            "kurztext": str(kurztext) if kurztext else "",
                            "langtext": str(langtext) if langtext else "",
                            "menge": menge if menge else 1.0,
                            "einheit": str(einheit) if einheit else "Psch"
                        }

                        positions.append(position)
                        print(f"[OK] Row {row}: {position['kurztext'][:50]}")

        workbook.close()
        print(f"\n📊 Total positions extracted: {len(positions)}")
        return positions

    except Exception as e:
        print(f"[ERROR] Error extracting from Excel: {e}")
        import traceback
        print(traceback.format_exc())
        return []

def estimate_prices_with_ai(positions, progress_callback=None):
    """
    Use AI to estimate prices for positions based on Langtext descriptions.
    Processes in batches to avoid token limits and JSON truncation.

    Args:
        positions: List of position dictionaries
        progress_callback: Optional function(percent, message) to report progress
    """
    # Helper function to normalize position numbers for comparison
    def normalize_pos(pos_str):
        """Normalize position number: remove trailing dots, leading zeros in segments"""
        if not pos_str:
            return ""
        pos_str = pos_str.rstrip('.')
        segments = pos_str.split('.')
        normalized = []
        for seg in segments:
            normalized.append(seg.lstrip('0') or '0')
        return '.'.join(normalized)

    def get_ai_prices_for_batch(batch_positions):
        """Get prices for a batch of positions from AI"""
        positions_text = "\n\n".join([
            f"Position:\n"
            f"Nummer: {pos['ordnungszahl']}\n"
            f"Beschreibung: {(pos['langtext'] or pos['kurztext'])[:500]}\n"
            f"Menge: {pos['menge']} {pos['einheit']}"
            for pos in batch_positions
        ])

        prompt = f"""Du bist ein erfahrener Baukalkulator. Gib für JEDE Position den EINHEITSPREIS (EP) in EUR.

⚠️ KRITISCH - EINHEITSPREIS (EP):
- Gib NUR den EINHEITSPREIS pro Einheit zurück!
- Bei "140 St" → Preis für 1 Stück (z.B. 150 EUR/St)
- Bei "1 psch" → Pauschalpreis für die gesamte Leistung
- Bei "psch" mit Mengenangaben im Text (z.B. "31.200 MeterWochen") → Multipliziere!

TYPISCHE EINHEITSPREISE:
- Verkehrszeichen (St): 100-200 EUR/St
- Baustelleneinrichtung (psch): 5.000-15.000 EUR
- Erdaushub (m³): 12-18 EUR/m³
- Asphalt (m²): 25-45 EUR/m²
- Beton C25/30 (m³): 180-260 EUR/m³
- Bauzaun mit MeterWochen (psch): Berechne aus Mengenangabe!

Positionen:
{positions_text}

Ausgabe NUR als JSON-Array:
[{{"pos": "Nummer", "unit_price": Preis}}, ...]
"""
        response, model_used = call_ai_with_retry(
            model='gemini-2.0-flash-lite',
            contents=[prompt]
        )

        text = response.text.strip()
        if text.startswith('```'):
            text = text.split('```')[1]
            if text.startswith('json'):
                text = text[4:]
            text = text.strip()

        return json.loads(text)

    def apply_prices_from_data(prices_data, target_positions):
        """Apply prices from AI response to positions"""
        price_lookup = {}
        for item in prices_data:
            pos_key = None
            for key in ['pos', 'nummer', 'position', 'ordnungszahl', 'nr']:
                if key in item:
                    pos_key = str(item[key])
                    break

            if pos_key:
                for key in ['unit_price', 'unitprice', 'preis', 'ep', 'einheitspreis', 'price']:
                    if key in item:
                        try:
                            price_lookup[pos_key] = float(item[key])
                        except:
                            pass
                        break
            else:
                for k, v in item.items():
                    try:
                        price_lookup[str(k)] = float(v)
                    except:
                        pass

        # Create normalized lookup
        normalized_lookup = {}
        for key, value in price_lookup.items():
            norm_key = normalize_pos(key)
            normalized_lookup[norm_key] = value
            if key not in normalized_lookup:
                normalized_lookup[key] = value

        matched = 0
        for pos in target_positions:
            pos_num = pos['ordnungszahl']
            norm_pos_num = normalize_pos(pos_num)

            if pos_num in price_lookup and price_lookup[pos_num] > 0:
                pos['unit_price'] = price_lookup[pos_num]
                matched += 1
            elif norm_pos_num in normalized_lookup and normalized_lookup[norm_pos_num] > 0:
                pos['unit_price'] = normalized_lookup[norm_pos_num]
                matched += 1

        return matched

    try:
        print(f"\n💰 Estimating prices with AI for {len(positions)} positions...")

        # Process in batches of 50 to avoid token limits
        BATCH_SIZE = 50
        total_matched = 0
        failed_batches = []

        # Progress goes from 30% to 85% during batch processing
        START_PROGRESS = 30
        END_PROGRESS = 85

        total_batches = (len(positions) + BATCH_SIZE - 1) // BATCH_SIZE

        for batch_start in range(0, len(positions), BATCH_SIZE):
            batch_end = min(batch_start + BATCH_SIZE, len(positions))
            batch = positions[batch_start:batch_end]
            batch_num = (batch_start // BATCH_SIZE) + 1

            # Calculate progress percentage for this batch
            progress_pct = START_PROGRESS + int((batch_num / total_batches) * (END_PROGRESS - START_PROGRESS))
            progress_msg = f"Schätze Preise... Batch {batch_num}/{total_batches}"

            if progress_callback:
                progress_callback(progress_pct, progress_msg)

            print(f"   📦 Processing batch {batch_num}/{total_batches} ({len(batch)} positions)...")

            try:
                prices_data = get_ai_prices_for_batch(batch)
                matched = apply_prices_from_data(prices_data, batch)
                total_matched += matched
                print(f"   ✓ Batch {batch_num}: {matched}/{len(batch)} prices matched")
            except Exception as batch_error:
                print(f"   ⚠️ Batch {batch_num} failed: {batch_error}")
                failed_batches.append((batch_num, batch))

            # Small delay between batches to avoid rate limits
            if batch_num < total_batches:
                time.sleep(2)

        # Retry failed batches after a longer delay
        if failed_batches:
            if progress_callback:
                progress_callback(86, f"Wiederhole {len(failed_batches)} fehlgeschlagene Batches...")
            print(f"   🔄 Retrying {len(failed_batches)} failed batches after delay...")
            time.sleep(10)
            for batch_num, batch in failed_batches:
                print(f"   📦 Retrying batch {batch_num}...")
                try:
                    prices_data = get_ai_prices_for_batch(batch)
                    matched = apply_prices_from_data(prices_data, batch)
                    total_matched += matched
                    print(f"   ✓ Batch {batch_num} retry: {matched}/{len(batch)} prices matched")
                except Exception as retry_error:
                    print(f"   ⚠️ Batch {batch_num} retry failed: {retry_error}")
                time.sleep(3)

        print(f"   📊 Total matched: {total_matched}/{len(positions)}")

        # Collect positions without prices for a second AI call
        missing_positions = []
        for pos in positions:
            current_price = pos.get('unit_price', 0)
            if current_price is None or current_price == 0:
                missing_positions.append(pos)

        # If there are missing positions, make a second AI call specifically for them
        if missing_positions:
            print(f"   {len(missing_positions)} positions without prices, making second AI call...")

            # Create focused prompt for missing positions only
            missing_text = "\n".join([
                f"Pos {p['ordnungszahl']}: {p.get('beschreibung', 'Keine Beschreibung')[:200]} | Einheit: {p.get('einheit', 'Psch')} | Menge: {p.get('menge', 1)}"
                for p in missing_positions
            ])

            fallback_prompt = f"""Du bist ein erfahrener deutscher Baukalkulant.
Gib realistische Einheitspreise (EP) in EUR für diese Baupositionen.

WICHTIG:
- Analysiere jede Beschreibung genau
- Gib realistische deutsche Baumarktpreise
- Alle Preise mit Cent-Beträgen (z.B. 125.50, nicht 125.00)

Positionen:
{missing_text}

Ausgabe NUR als JSON-Array:
[
  {{"pos": "Nummer", "unit_price": Preis}},
  ...
]
"""
            try:
                fallback_response, _ = call_ai_with_retry(
                    model='gemini-2.0-flash',  # Use slightly better model for retry
                    contents=[fallback_prompt]
                )

                fallback_text = fallback_response.text.strip()
                if fallback_text.startswith('```'):
                    fallback_text = fallback_text.split('```')[1]
                    if fallback_text.startswith('json'):
                        fallback_text = fallback_text[4:]
                    fallback_text = fallback_text.strip()

                fallback_prices = json.loads(fallback_text)
                print(f"   Second AI call returned {len(fallback_prices)} prices")

                # Build lookup from fallback response
                fallback_lookup = {}
                for item in fallback_prices:
                    pos_key = None
                    for key in ['pos', 'nummer', 'position', 'ordnungszahl', 'nr']:
                        if key in item:
                            pos_key = str(item[key])
                            break

                    if pos_key:
                        for key in ['unit_price', 'unitprice', 'preis', 'ep', 'einheitspreis', 'price']:
                            if key in item:
                                try:
                                    fallback_lookup[pos_key] = float(item[key])
                                except:
                                    pass
                                break
                    else:
                        for k, v in item.items():
                            try:
                                fallback_lookup[str(k)] = float(v)
                            except:
                                pass

                # Create normalized fallback lookup
                normalized_fallback = {}
                for key, value in fallback_lookup.items():
                    norm_key = normalize_pos(key)
                    normalized_fallback[norm_key] = value
                    if key not in normalized_fallback:
                        normalized_fallback[key] = value

                # Apply fallback prices from second AI call
                for pos in missing_positions:
                    pos_num = pos['ordnungszahl']
                    norm_pos_num = normalize_pos(pos_num)

                    if pos_num in fallback_lookup and fallback_lookup[pos_num] > 0:
                        pos['unit_price'] = fallback_lookup[pos_num]
                        print(f"   AI fallback matched {pos_num}: {pos['unit_price']}")
                    elif norm_pos_num in normalized_fallback and normalized_fallback[norm_pos_num] > 0:
                        pos['unit_price'] = normalized_fallback[norm_pos_num]
                        print(f"   AI fallback matched {pos_num} (normalized): {pos['unit_price']}")

            except Exception as fallback_error:
                print(f"   Second AI call failed: {fallback_error}")

        # Final fallback: use unit-based defaults for any still missing
        final_missing = 0
        for pos in positions:
            current_price = pos.get('unit_price', 0)
            if current_price is None or current_price == 0:
                final_missing += 1
                unit = pos.get('einheit', 'Psch').lower()
                if 'psch' in unit or 'pau' in unit:
                    pos['unit_price'] = 1500.50
                elif 'm³' in unit or 'm3' in unit:
                    pos['unit_price'] = 125.75
                elif 'm²' in unit or 'm2' in unit:
                    pos['unit_price'] = 65.50
                elif 'm' in unit:
                    pos['unit_price'] = 35.25
                else:
                    pos['unit_price'] = 85.50
                print(f"   Final fallback for {pos['ordnungszahl']}: {pos['unit_price']}")

        if final_missing > 0:
            print(f"   ⚠️ {final_missing} positions still needed unit-based fallback")

        print(f"✓ Price estimation complete")
        return positions
    
    except Exception as e:
        print(f"⚠️ Error estimating prices: {e}")
        # Add default prices
        for pos in positions:
            if 'unit_price' not in pos or pos['unit_price'] == 0:
                pos['unit_price'] = 100.50
        return positions

def read_excel_as_text(file_path):
    """
    Read Excel file and convert to text format for AI processing.
    """
    try:
        # Read all sheets with context manager to ensure proper cleanup
        with pd.ExcelFile(file_path) as excel_file:
            text_content = ""

            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                text_content += f"\n\n=== SHEET: {sheet_name} ===\n\n"
                text_content += df.to_string(index=False)

        return text_content
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

def read_excel_as_text_chunked(file_path, max_rows=500):
    """
    Read Excel file with row limit to prevent token overflow.
    Only reads up to max_rows per sheet.
    """
    try:
        with pd.ExcelFile(file_path) as excel_file:
            text_content = ""

            for sheet_name in excel_file.sheet_names:
                # Read with row limit
                df = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=max_rows)

                text_content += f"\n\n=== SHEET: {sheet_name} ===\n\n"
                text_content += df.to_string(index=False)

                # Add note if we truncated
                total_rows = pd.read_excel(excel_file, sheet_name=sheet_name, usecols=[0]).shape[0]
                if total_rows > max_rows:
                    text_content += f"\n\n[NOTE: Showing first {max_rows} of {total_rows} rows to prevent token limit]"

        return text_content
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

def extract_with_ai(file_path, file_extension, progress_bar=None, status_text=None):
    """
    Master extraction function - sends file directly to AI for complete analysis.

    Args:
        file_path: Path to the file to process
        file_extension: File extension (e.g., '.pdf', '.xlsx')
        progress_bar: Optional Streamlit progress bar to update
        status_text: Optional Streamlit text element to update status
    """
    def update_progress(percent, message):
        """Helper to update progress bar and status text"""
        if progress_bar is not None:
            progress_bar.progress(percent / 100, text=f"{percent}% - {message}")
        if status_text is not None:
            status_text.text(message)
        print(f"[{percent}%] {message}")

    # Start total timer
    total_start_time = time.time()

    try:
        update_progress(5, "Starte Analyse...")

        print(f"\n{'='*70}")
        print(f"🤖 AI-POWERED EXTRACTION")
        print(f"📄 File: {os.path.basename(file_path)}")
        print(f"📝 Extension: {file_extension}")
        print(f"{'='*70}\n")

        # Check if file is Excel - handle differently
        ext = file_extension.lower()
        if ext in ['.xlsx', '.xls']:
            # First check if Excel has the expected structure
            update_progress(10, "Prüfe Excel-Struktur...")
            print(f"🔍 Checking Excel structure...")
            has_structure = check_excel_structure(file_path)

            if has_structure:
                update_progress(15, "Excel-Struktur erkannt - Extrahiere Positionen...")
                print(f"✓ Excel has expected structure - using direct extraction")
                print(f"📊 Extracting positions from Excel...")

                # Extract positions
                positions = extract_positions_from_structured_excel(file_path)

                if positions:
                    # Estimate prices with AI (pass progress callback)
                    update_progress(30, "Schätze Preise mit KI...")
                    positions = estimate_prices_with_ai(positions, progress_callback=update_progress)

                    # Convert to DataFrame
                    data = []
                    for pos in positions:
                        try:
                            qty = float(pos.get('menge', 1.0)) if pos.get('menge') else 1.0
                        except:
                            qty = 1.0

                        data.append({
                            'pos': pos.get('ordnungszahl', ''),
                            'description': pos.get('langtext', pos.get('kurztext', '')),
                            'quantity': qty,
                            'unit': pos.get('einheit', 'Psch'),
                            'unit_price': pos.get('unit_price', 0.0)
                        })

                    update_progress(90, "Erstelle Ergebnis-Tabelle...")
                    df = pd.DataFrame(data)
                    total_time = time.time() - total_start_time
                    update_progress(100, f"Fertig! {len(df)} Positionen extrahiert")
                    print(f"\n✅ Extraction complete in {total_time:.2f}s")
                    print(f"📊 Extracted {len(df)} positions")
                    return df
                else:
                    print(f"⚠️ No positions found - falling back to AI extraction")
            else:
                print(f"⚠️ Excel structure doesn't match - using AI extraction")

            # Fallback: Convert Excel to text (chunked to avoid token limits)
            # Gemini doesn't support Excel MIME type directly
            update_progress(20, "Konvertiere Excel zu Text...")
            print(f"📊 Processing Excel file with AI (text conversion)...")
            read_start = time.time()

            # Read Excel with size limit to prevent token overflow
            excel_text = read_excel_as_text_chunked(file_path, max_rows=500)
            read_time = time.time() - read_start
            update_progress(30, "Excel-Datei gelesen")
            print(f"✅ Excel file read successfully ({read_time:.2f}s)")

            if excel_text is None:
                raise Exception("Failed to read Excel file")

            # Check text size and warn if too large
            text_length = len(excel_text)
            estimated_tokens = text_length // 4  # Rough estimate: 1 token ≈ 4 characters
            print(f"   Text size: {text_length:,} characters (~{estimated_tokens:,} tokens)")

            if estimated_tokens > 900000:  # Leave margin below 1M token limit
                print(f"⚠️ File is very large, using only first 500 rows to avoid token limit")

            # Send text content to AI
            update_progress(40, "KI analysiert Dokument...")
            print(f"\n🧠 AI analyzing document...")
            print(f"   Starting with: gemini-2.5-flash-lite (will auto-switch if needed)")

            analysis_start = time.time()
            prompt_with_data = f"{MASTER_EXTRACTION_PROMPT}\n\nDOKUMENT INHALT:\n{excel_text}"
            response, model_used = call_ai_with_retry(
                model='gemini-2.5-flash-lite',
                contents=[prompt_with_data]
            )
        else:
            # For other file types, upload to Gemini
            update_progress(15, "Lade Datei zur KI hoch...")
            print(f"📤 Uploading file to AI...")
            upload_start = time.time()
            
            # Determine MIME type
            mime_type = get_mime_type(file_path)
            print(f"   MIME Type: {mime_type}")
            
            # Upload file to Gemini
            try:
                # Try with mime_type parameter
                file_ref = genai.upload_file(
                    path=file_path,
                    mime_type=mime_type
                )
            except TypeError:
                # Fallback: let it auto-detect
                file_ref = genai.upload_file(path=file_path)
            
            upload_time = time.time() - upload_start
            update_progress(30, "Datei hochgeladen")
            print(f"✅ File uploaded successfully ({upload_time:.2f}s)")
            print(f"   File URI: {file_ref.uri}")
            print(f"   File Name: {file_ref.name}")

            # Send to AI with master prompt
            update_progress(40, "KI analysiert Dokument...")
            print(f"\n🧠 AI analyzing document...")
            print(f"   Starting with: gemini-2.5-flash-lite (will auto-switch if needed)")

            analysis_start = time.time()
            response, model_used = call_ai_with_retry(
                model='gemini-2.5-flash-lite',
                contents=[file_ref, MASTER_EXTRACTION_PROMPT]
            )
        analysis_time = time.time() - analysis_start
        update_progress(70, "KI-Antwort erhalten")

        if model_used != 'gemini-2.5-flash-lite':
            print(f"   ✓ Used model: {model_used}")

        text = response.text
        print(f"\n📥 AI Response received ({analysis_time:.2f}s)")
        print(f"   Length: {len(text)} characters")

        # Parse JSON response
        update_progress(80, "Verarbeite KI-Antwort...")
        parse_start = time.time()
        df = parse_json_response(text)
        parse_time = time.time() - parse_start
        print(f"   Parsing time: {parse_time:.2f}s")
        
        if not df.empty:
            print(f"\n✅ Extraction successful: {len(df)} positions found")

            # Check for zero prices
            zero_prices = (df['unit_price'] == 0).sum()
            if zero_prices > 0:
                update_progress(85, f"Korrigiere {zero_prices} Preise mit KI...")
                print(f"⚠️  Warning: {zero_prices} positions with zero price")
                print(f"🔧 Requesting AI to fix prices...")
                price_fix_start = time.time()
                df = fix_prices_with_ai(df)
                price_fix_time = time.time() - price_fix_start
                print(f"   Price fixing time: {price_fix_time:.2f}s")

            # Calculate total time
            total_time = time.time() - total_start_time
            update_progress(100, f"Fertig! {len(df)} Positionen extrahiert")
            print(f"\n{'='*70}")
            print(f"⏱️  TOTAL PROCESSING TIME: {total_time:.2f} seconds ({total_time/60:.2f} minutes)")
            print(f"{'='*70}\n")
            
            return df
        else:
            print(f"❌ No positions extracted from AI response")
            total_time = time.time() - total_start_time
            print(f"⏱️  Total time: {total_time:.2f}s")
            return pd.DataFrame(columns=["pos", "description", "quantity", "unit", "unit_price"])
    
    except Exception as e:
        print(f"❌ AI Extraction Error: {str(e)}")
        print(f"🔍 Details: {traceback.format_exc()}")
        total_time = time.time() - total_start_time
        print(f"⏱️  Time before error: {total_time:.2f}s")
        return pd.DataFrame(columns=["pos", "description", "quantity", "unit", "unit_price"])

def parse_json_response(text):
    """
    Parse JSON from AI response with multiple fallback strategies.
    """
    try:
        # Clean the response
        text = text.strip()
        
        print(f"🔍 Parsing JSON response...")
        
        # Strategy 1: Direct JSON parse
        try:
            data = json.loads(text)
            if isinstance(data, list) and len(data) > 0:
                df = pd.DataFrame(data)
                print(f"✓ JSON parsed (direct): {len(df)} positions")
                return df
        except:
            pass
        
        # Strategy 2: Extract from markdown code blocks
        patterns = [
            r'```json\s*(\[.*?\])\s*```',  # ```json [...] ```
            r'```\s*(\[.*?\])\s*```',       # ``` [...] ```
            r'(\[.*?\])',                    # [...] anywhere in text
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, text, re.DOTALL)
            for match in matches:
                try:
                    data = json.loads(match)
                    if isinstance(data, list) and len(data) > 0:
                        df = pd.DataFrame(data)
                        print(f"✓ JSON parsed (pattern match): {len(df)} positions")
                        return df
                except:
                    continue
        
        # Strategy 3: Find JSON array manually
        start = text.find('[')
        end = text.rfind(']')
        if start != -1 and end != -1 and end > start:
            try:
                json_str = text[start:end+1]
                data = json.loads(json_str)
                if isinstance(data, list) and len(data) > 0:
                    df = pd.DataFrame(data)
                    print(f"✓ JSON parsed (manual extraction): {len(df)} positions")
                    return df
            except:
                pass
        
        print(f"⚠️  Could not parse JSON from response")
        print(f"📄 Response preview: {text[:500]}")
        return pd.DataFrame(columns=["pos", "description", "quantity", "unit", "unit_price"])
    
    except Exception as e:
        print(f"❌ JSON parsing error: {e}")
        return pd.DataFrame(columns=["pos", "description", "quantity", "unit", "unit_price"])

def fix_prices_with_ai(df):
    """
    Send positions back to AI to fix zero or missing prices.
    """
    try:
        print(f"\n🔧 PRICE CORRECTION WITH AI")
        
        # Convert to JSON
        positions_json = df.to_json(orient='records', force_ascii=False, indent=2)
        
        print(f"📤 Sending {len(df)} positions for price correction...")
        
        # Create prompt with positions
        prompt = PRICING_CORRECTION_PROMPT.format(positions_json=positions_json)
        
        # Send to AI with retry logic
        response, model_used = call_ai_with_retry(
            model='gemini-2.5-flash-lite',
            contents=[prompt]
        )
        
        if model_used != 'gemini-2.5-flash-lite':
            print(f"   ✓ Used alternate model: {model_used}")
        
        # Parse response
        df_fixed = parse_json_response(response.text)
        
        if not df_fixed.empty:
            zero_after = (df_fixed['unit_price'] == 0).sum()
            print(f"✅ Prices fixed: {zero_after} zero prices remaining")
            return df_fixed
        else:
            print(f"⚠️  Price fix failed, returning original")
            return df
            
    except Exception as e:
        print(f"❌ Price fixing error: {e}")
        return df

# --- PDF GENERATION ---
class OfferPDF(FPDF):
    def header(self):
        self.set_font("Arial", 'B', 14)
        self.set_text_color(0, 0, 0)
        self.cell(100, 8, "RUTTENSCHEID BAUKONZEPTE", ln=0, align='L')
        self.set_font("Arial", '', 10)
        self.cell(0, 8, f"Datum: {datetime.now().strftime('%d.%m.%Y')}", ln=1, align='R')
        self.set_font("Arial", '', 10)
        self.cell(100, 5, "Munchener Str. 100 A, 45145 Essen", ln=1, align='L')
        self.ln(10)
    
    def footer(self):
        self.set_y(-35)
        self.set_draw_color(200, 200, 200)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(2)
        self.set_font("Arial", '', 8)
        self.set_text_color(80, 80, 80)
        
        col_width = 63
        x_start = 10
        self.set_xy(x_start, self.get_y())
        self.multi_cell(col_width, 4,
            "Ruttenscheid Baukonzepte GmbH\n"
            "Munchener Str. 100A\n"
            "45145 Essen\n"
            "Geschaftsfuhrer: Dipl.-Ing. Moh Alturky", align='L')
        
        self.set_xy(x_start + col_width, self.get_y() - 16)
        self.multi_cell(col_width, 4,
            "Tel: +49 0201 84850166\n"
            "Mob: +49 160 7901911\n"
            "E-Mail: Moh@ruttenscheid-bau.de\n"
            "Web: www.ruttenscheid-bau.de", align='C')
        
        self.set_xy(x_start + (col_width * 2), self.get_y() - 16)
        self.cell(col_width, 4, f"Seite {self.page_no()}", align='R')

def generate_offer_pdf(df, project_name):
    """Generate professional PDF offer."""
    pdf = OfferPDF()
    pdf.set_auto_page_break(auto=True, margin=40)
    pdf.add_page()
    
    def clean(text):
        if pd.isna(text):
            return ""
        text = str(text)
        replacements = {
            "€": "EUR", "–": "-", "„": '"', "“": '"', "'": "'",
            "ä": "ae", "ö": "oe", "ü": "ue", "ß": "ss", "Ä": "Ae",
            "Ö": "Oe", "Ü": "Ue", "²": "2", "³": "3"
        }
        for char, rep in replacements.items():
            text = text.replace(char, rep)
        return text.encode('latin-1', 'replace').decode('latin-1')
    
    pdf.ln(5)
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(100, 8, "Auftraggeber", 0, 0, 'L')
    pdf.cell(90, 8, clean(f"Angebot Nr. {datetime.now().strftime('%Y-%m-%d')}"), 0, 1, 'R')
    pdf.set_font("Arial", '', 11)
    pdf.cell(100, 6, clean("Bauherr"), 0, 0, 'L')
    pdf.cell(90, 6, clean(f"Projekt: {project_name}"), 0, 1, 'R')
    pdf.ln(15)
    
    pdf.set_font("Arial", 'B', 14)
    pdf.cell(0, 10, clean(f"Angebot: {project_name}"), ln=1, align='L')
    pdf.set_font("Arial", '', 11)
    pdf.multi_cell(0, 6, clean("Sehr geehrte Damen und Herren,\nhiermit unterbreiten wir Ihnen unser Angebot gemaß Ihrer Anfrage."))
    pdf.ln(10)
    
    # Table Header
    pdf.set_fill_color(230, 230, 230)
    pdf.set_font("Arial", 'B', 9)
    pdf.set_draw_color(180, 180, 180)
    
    w = [20, 85, 20, 15, 25, 25]
    pdf.cell(w[0], 8, "Pos.", 1, 0, 'C', 1)
    pdf.cell(w[1], 8, "Bezeichnung", 1, 0, 'L', 1)
    pdf.cell(w[2], 8, "Menge", 1, 0, 'C', 1)
    pdf.cell(w[3], 8, "Einh.", 1, 0, 'C', 1)
    pdf.cell(w[4], 8, "EP (EUR)", 1, 0, 'R', 1)
    pdf.cell(w[5], 8, "GP (EUR)", 1, 1, 'R', 1)
    
    # Table Content
    pdf.set_font("Arial", size=9)
    total_netto = 0.0
    
    for _, row in df.iterrows():
        try:
            qty = float(row.get('quantity', 0))
            ep = float(row.get('unit_price', 0))
        except:
            qty, ep = 0, 0
        
        gp = qty * ep
        total_netto += gp
        
        pos = clean(str(row.get('pos', '')))
        desc = clean(str(row.get('description', '')))[:45]
        unit = clean(str(row.get('unit', '')))
        
        pdf.cell(w[0], 8, pos, 1, 0, 'C')
        pdf.cell(w[1], 8, desc, 1, 0, 'L')
        pdf.cell(w[2], 8, f"{qty:.2f}".replace('.', ','), 1, 0, 'C')
        pdf.cell(w[3], 8, unit, 1, 0, 'C')
        pdf.cell(w[4], 8, format_german_number(ep), 1, 0, 'R')
        pdf.cell(w[5], 8, format_german_number(gp), 1, 1, 'R')
    
    # Totals
    pdf.ln(5)
    
    def print_total(label, value, bold=False):
        pdf.set_font("Arial", 'B' if bold else '', 10)
        pdf.cell(145, 8, clean(label), 0, 0, 'R')
        pdf.cell(45, 8, f"{format_german_number(value)} EUR", 1 if bold else 0, 1, 'R')
    
    print_total("Summe Netto:", total_netto)
    print_total("zzgl. 19% MwSt.:", total_netto * 0.19)
    pdf.ln(2)
    print_total("Gesamtbetrag (Brutto):", total_netto * 1.19, bold=True)
    
    pdf.ln(15)
    pdf.set_font("Arial", '', 10)
    pdf.multi_cell(0, 5, clean("Wir hoffen, Ihnen ein interessantes Angebot unterbreitet zu haben und stehen fur Ruckfragen gerne zur Verfugung."))
    pdf.ln(10)
    pdf.cell(0, 10, clean("Mit freundlichen Grußen"), ln=1)
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(0, 10, clean("Ruttenscheid Baukonzepte GmbH"), ln=1)
    
    return pdf.output(dest='S').encode('latin-1', 'replace')


# --- STREAMLIT UI ---
st.set_page_config(
    page_title="Rüttenscheid Smart Kalkulation",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Initialize API Client
# Load API key from Streamlit secrets (for deployment) or environment variable (for local)
try:
    api_key = st.secrets.get("GEMINI_API_KEY") or os.environ.get("GEMINI_API_KEY")
except:
    api_key = os.environ.get("GEMINI_API_KEY")

if not api_key:
    st.error("⚠️ API Key fehlt! Bitte in Streamlit Secrets konfigurieren.")
    st.info("💡 Für Deployment: Fügen Sie GEMINI_API_KEY in den App-Secrets hinzu.")
    st.stop()

genai.configure(api_key=api_key)

# Header with logo and company name centered
logo_path = "Data/Screenshot 2026-01-07 214122.png"

# Create centered header with logo and text
col1, col_center, col3 = st.columns([1, 2, 1])

with col_center:
    if os.path.exists(logo_path):
        # Center the smaller logo using columns
        logo_col1, logo_col2, logo_col3 = st.columns([1, 1, 1])
        with logo_col2:
            st.image(logo_path, width=200)
    
    st.markdown(f"""
    <div style='text-align: center;'>
        <h1 style='color: #1e3a8a; margin: 1px0 0; font-size: 2.5em;'>{COMPANY_NAME}</h1>
        <p style='color: #3b82f6; font-size: 1.2em; margin: 10px 0 0 0;'>🤖 KI-Gestützte Angebotserstellung</p>
    </div>
    """, unsafe_allow_html=True)


# Initialize session state
if "calculation_df" not in st.session_state:
    st.session_state.calculation_df = pd.DataFrame(
        columns=["pos", "description", "quantity", "unit", "original_price", "unit_price", "price_factor"]
    )
if "price_factor" not in st.session_state:
    st.session_state.price_factor = 1.0
if "project_name" not in st.session_state:
    st.session_state.project_name = "Bauprojekt"
if "project_link" not in st.session_state:
    st.session_state.project_link = ""
if "file_uploader_key" not in st.session_state:
    st.session_state.file_uploader_key = 0
if "folder_location" not in st.session_state:
    st.session_state.folder_location = os.path.expanduser("~\\Desktop")

# Helper function for folder path input (cloud-compatible)
def select_folder():
    """In cloud environment, folder selection not needed - files are downloaded directly."""
    return False

def is_cloud_environment():
    """Check if running in Streamlit Cloud"""
    return not os.path.exists(os.path.expanduser("~\\Desktop"))

# Step 1: Upload
st.markdown("---")
col_header1, col_header2 = st.columns([4, 1])
with col_header1:
    st.subheader("📤 Schritt 1: Dokument hochladen")
with col_header2:
    if not st.session_state.calculation_df.empty:
        if st.button("🔄 Neu starten", use_container_width=True, help="Alle Daten löschen und von vorne beginnen"):
            st.session_state.calculation_df = pd.DataFrame(
                columns=["pos", "description", "quantity", "unit", "original_price", "unit_price", "price_factor"]
            )
            st.session_state.price_factor = 1.0
            st.session_state.file_uploader_key += 1  # Reset file uploader
            st.rerun()

uploaded_file = st.file_uploader(
    "Wählen Sie eine Datei:",
    type=['pdf', 'docx', 'doc', 'txt', 'xlsx', 'xls',
          'd81', 'd82', 'd83', 'd84', 'd85', 'd86', 'd90',
          'x81', 'x82', 'x83', 'x84', 'x85', 'x86', 'x90',
          'p81', 'p82', 'p83', 'p84', 'p85', 'p86', 'p90'],
    help="Unterstützte Formate: GAEB (D/X/P), PDF, Word, Excel",
    key=f"file_uploader_{st.session_state.file_uploader_key}"
)

if uploaded_file:
    file_ext = uploaded_file.name.split('.')[-1].upper()
    file_size = uploaded_file.size / 1024
    
    st.markdown("""<div style='background-color: #f0f9ff; padding: 15px; border-radius: 8px; border-left: 4px solid #3b82f6; margin: 10px 0;'>""", unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f"**📄 Datei:**<br>{uploaded_file.name}", unsafe_allow_html=True)
    with col2:
        st.markdown(f"**📋 Format:**<br>{file_ext}", unsafe_allow_html=True)
    with col3:
        st.markdown(f"**💾 Größe:**<br>{file_size:.1f} KB", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)
    
    if st.button("🚀 Jetzt analysieren", type="primary", use_container_width=True, help="Dokument mit KI analysieren und Positionen extrahieren"):
        # Create progress bar and status text
        progress_bar = st.progress(0, text="0% - Starte...")
        status_text = st.empty()

        # Save uploaded file temporarily
        suffix = f".{uploaded_file.name.split('.')[-1]}"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(uploaded_file.getvalue())
            temp_path = tmp.name

        try:
            # Extract with AI (pass progress bar)
            df_result = extract_with_ai(temp_path, suffix.lower(), progress_bar=progress_bar, status_text=status_text)

            # Cleanup temp file
            safe_remove_file(temp_path)

            # Clear progress bar after completion
            progress_bar.empty()
            status_text.empty()

            if not df_result.empty:
                # Clean and validate data
                df_result = df_result[df_result['description'].notna()]
                df_result = df_result[df_result['description'].str.len() > 3]
                df_result['quantity'] = pd.to_numeric(df_result['quantity'], errors='coerce').fillna(1.0)
                df_result['unit_price'] = pd.to_numeric(df_result['unit_price'], errors='coerce').fillna(0.0)
                # Store original prices
                df_result['original_price'] = df_result['unit_price'].copy()
                # Initialize price factor
                df_result['price_factor'] = 1.0
                df_result = df_result.reset_index(drop=True)

                st.session_state.calculation_df = df_result
                st.session_state.price_factor = 1.0
                st.success(f"✅ **Erfolgreich!** {len(df_result)} Positionen extrahiert")

                # Statistics with enhanced display
                st.markdown("#### 📊 Extraktionsergebnis")
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("📋 Positionen", f"{len(df_result)}", help="Anzahl der gefundenen Positionen")
                with col2:
                    priced = (df_result['unit_price'] > 0).sum()
                    st.metric("💰 Mit Preis", f"{priced}", help="Positionen mit Preisangabe")
                with col3:
                    total = (df_result['quantity'] * df_result['unit_price']).sum()
                    st.metric("💵 Summe Netto", f"{format_german_number(total, 0)} €", help="Gesamtsumme ohne MwSt.")
                with col4:
                    total_brutto = total * 1.19
                    st.metric("✅ Summe Brutto", f"{format_german_number(total_brutto, 0)} €", help="Gesamtsumme inkl. 19% MwSt.")
            else:
                st.error("❌ Keine Positionen gefunden. Bitte prüfen Sie das Dokument.")

        except Exception as e:
            # Clear progress bar on error
            progress_bar.empty()
            status_text.empty()
            st.error(f"❌ Fehler: {str(e)}")
            if "503" in str(e) or "overloaded" in str(e).lower():
                st.warning("⚠️ Alle verfügbaren KI-Modelle sind derzeit überlastet. Bitte versuchen Sie es in einigen Minuten erneut.")
                st.info("💡 Das System hat automatisch folgende Modelle versucht: gemini-2.5-flash, gemini-2.5-flash-lite, gemini-2.0-flash, gemini-2.0-flash-lite, gemini-2.5-pro")
            elif "429" in str(e) or "quota" in str(e).lower() or "RESOURCE_EXHAUSTED" in str(e):
                if "input_token" in str(e) or "token" in str(e).lower():
                    st.warning("⚠️ Token-Limit für alle verfügbaren Modelle überschritten. Das System hat bereits mehrere Modelle versucht.")
                    st.info("💡 Tipp: Bei sehr großen Dateien kann es zu Token-Limits kommen. Versuchen Sie kleinere Dateien oder warten Sie 1-2 Minuten.")
                else:
                    st.warning("⚠️ API-Quota für alle verfügbaren Modelle überschritten.")
                    st.info("💡 Das System hat automatisch 5 verschiedene Modelle versucht. Bitte warten Sie einige Minuten.")
            elif "Unsupported MIME type" in str(e) or "INVALID_ARGUMENT" in str(e):
                st.warning("⚠️ Dieses Dateiformat wird möglicherweise nicht direkt unterstützt. Das System verarbeitet das Dokument lokal.")
            safe_remove_file(temp_path)

st.markdown("---")

# Step 2: Edit & Calculate
st.markdown("---")
st.subheader("✏️ Schritt 2: Positionen bearbeiten")

if not st.session_state.calculation_df.empty:
    
    # Calculate total price for display (GP = Menge × EP)
    display_df = st.session_state.calculation_df.copy()
    # GP netto = quantity × unit_price (EP already includes any factor applied)
    display_df['total_price'] = display_df['quantity'] * display_df['unit_price']

    # Create formatted display columns for German number format
    display_df['quantity_display'] = display_df['quantity'].apply(lambda x: format_german_number(x, 2))
    display_df['unit_price_display'] = display_df['unit_price'].apply(lambda x: format_german_number(x, 2))
    display_df['total_price_display'] = display_df['total_price'].apply(lambda x: format_german_number(x, 2))

    edited_df = st.data_editor(
        display_df,
        column_config={
            "pos": st.column_config.TextColumn("Pos.", width="small"),
            "description": st.column_config.TextColumn("Leistungsbezeichnung", width="large"),
            "quantity_display": st.column_config.TextColumn("Menge", width="small", help="Menge im deutschen Format"),
            "unit": st.column_config.TextColumn("Einheit", width="small"),
            "unit_price_display": st.column_config.TextColumn("EP netto (€)", width="medium", help="Einheitspreis: 1.234,56 €"),
            "total_price_display": st.column_config.TextColumn("GP netto (€)", width="medium", disabled=True, help="Gesamtpreis (Menge × EP × Faktor): 3.602,40 €"),
            "quantity": None,  # Hide raw numeric columns
            "unit_price": None,
            "total_price": None,
            "original_price": None,
            "price_factor": None
        },
        column_order=["pos", "description", "quantity_display", "unit", "unit_price_display", "total_price_display"],
        num_rows="dynamic",
        use_container_width=True,
        height=400,
        key="editor",
        disabled=["total_price_display"]
    )

    # Convert German formatted strings back to numbers for calculations
    def parse_german_number(value_str):
        """Convert German formatted string back to float: '1.234,56' -> 1234.56"""
        if pd.isna(value_str) or value_str == "":
            return 0.0
        try:
            # Remove thousand separators (dots) and replace decimal comma with dot
            cleaned = str(value_str).replace('.', '').replace(',', '.')
            return float(cleaned)
        except:
            return 0.0

    # Update numeric values from edited German-formatted strings
    if 'quantity_display' in edited_df.columns:
        edited_df['quantity'] = edited_df['quantity_display'].apply(parse_german_number)
    if 'unit_price_display' in edited_df.columns:
        edited_df['unit_price'] = edited_df['unit_price_display'].apply(parse_german_number)
    
    # Recalculate total_price (GP = Menge × EP)
    edited_df['quantity'] = pd.to_numeric(edited_df['quantity'], errors='coerce').fillna(0)
    edited_df['unit_price'] = pd.to_numeric(edited_df['unit_price'], errors='coerce').fillna(0)
    edited_df['total_price'] = edited_df['quantity'] * edited_df['unit_price']
    
    # Check if data has changed (comparing key columns)
    old_df = st.session_state.calculation_df
    data_changed = False
    
    if len(edited_df) != len(old_df):
        data_changed = True
    else:
        # Compare quantities and unit prices
        for idx in range(len(edited_df)):
            if idx < len(old_df):
                if (edited_df.iloc[idx]['quantity'] != old_df.iloc[idx]['quantity'] or
                    edited_df.iloc[idx]['unit_price'] != old_df.iloc[idx]['unit_price']):
                    data_changed = True
                    break
    
    # Update session state with edited values
    st.session_state.calculation_df = edited_df
    
    # If data changed, trigger rerun to update display
    if data_changed:
        st.rerun()
    
    # Price multiplier with enhanced UI
    st.markdown("")
    st.markdown("**🔢 Preisanpassung - Alle Preise auf einmal ändern**")
    st.markdown("Passen Sie alle Einheitspreise gleichzeitig mit einem Faktor an.")
    col_mult1, col_mult2, col_mult3 = st.columns([2, 2, 1])
    with col_mult1:
        price_multiplier = st.number_input(
            "Multiplikator:",
            min_value=0.1,
            max_value=10.0,
            value=1.0,
            step=0.1,
            help="Faktor zur Preisanpassung"
        )
    with col_mult2:
        st.markdown("<div style='margin-top: 20px;'>", unsafe_allow_html=True)
        example_change = ((price_multiplier - 1) * 100)
        if price_multiplier > 1:
            st.info(f"➕ {example_change:.0f}% Aufschlag")
        elif price_multiplier < 1:
            st.info(f"➖ {abs(example_change):.0f}% Abschlag")
        else:
            st.info("Keine Änderung")
        st.markdown("</div>", unsafe_allow_html=True)
    with col_mult3:
        st.markdown("<div style='margin-top: 28px;'>", unsafe_allow_html=True)
        if st.button("✅ Anwenden", use_container_width=True, type="primary"):
            # Update unit_price directly (EP changes, GP recalculates automatically)
            st.session_state.calculation_df['unit_price'] = st.session_state.calculation_df['unit_price'] * price_multiplier
            st.session_state.price_factor = price_multiplier
            st.success(f"✅ Faktor {price_multiplier} angewendet! EP wurde angepasst.")
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)
    
    # Calculate totals (already calculated in edited_df above, just sum them)
    
    total_netto = edited_df["total_price"].sum()
    total_mwst = total_netto * 0.19
    total_brutto = total_netto * 1.19
    
    # Display totals
    st.markdown("")
    st.markdown("### 💰 Gesamtkalkulation")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("💵 Summe Netto", f"{format_german_number(total_netto)} €", help="Gesamtsumme ohne Mehrwertsteuer")
    with col2:
        st.metric("📊 MwSt. (19%)", f"{format_german_number(total_mwst)} €", help="Mehrwertsteuer 19%")
    with col3:
        st.metric("✅ Summe Brutto", f"{format_german_number(total_brutto)} €", delta=f"+{format_german_number(total_mwst)} €", help="Gesamtsumme inkl. MwSt.")
    
    st.markdown("---")
    
    # Folder Generator - Works in both environments
    with st.expander("📁 Projektordner-Generator", expanded=False):
        in_cloud = is_cloud_environment()

        if in_cloud:
            st.markdown("Erstellen Sie eine ZIP-Datei mit strukturierter Ordnerstruktur für Ihr Bauprojekt.")
        else:
            st.markdown("Erstellen Sie automatisch eine strukturierte Ordnerstruktur für Ihr Bauprojekt.")

        col_folder1, col_folder2 = st.columns(2)

        with col_folder1:
            # Only show folder location selector in local environment
            if not in_cloud:
                st.markdown("**📍 Speicherort:**")
                col_path, col_btn = st.columns([3, 1])
                with col_path:
                    st.text_input(
                        "Ausgewählter Pfad:",
                        value=st.session_state.folder_location,
                        disabled=True,
                        label_visibility="collapsed"
                    )
                with col_btn:
                    if st.button("📁", help="Ordner auswählen", key="select_folder_btn", use_container_width=True):
                        if select_folder():
                            st.success("✅ Ordner ausgewählt!")
                            st.rerun()
            else:
                st.markdown("**📦 Export als ZIP-Datei**")
                st.info("Die Ordnerstruktur wird als ZIP-Datei zum Download bereitgestellt.")

        with col_folder2:
            project_name_for_folder = st.text_input(
                "📝 Name des Hauptordners:",
                value=st.session_state.project_name if st.session_state.project_name else "Neues_Projekt",
                help="Geben Sie den Namen für den Hauptordner ein",
                key="folder_name_input",
                on_change=lambda: setattr(st.session_state, 'project_name', st.session_state.folder_name_input)
            )

        st.markdown("")
        project_link = st.text_input(
            "🔗 Projekt-Link (optional):",
            value=st.session_state.project_link if st.session_state.project_link else "",
            help="Geben Sie einen Link zum Projekt ein (z.B. Cloud-Speicher, Projektmanagement-Tool)",
            key="project_link_input",
            on_change=lambda: setattr(st.session_state, 'project_link', st.session_state.project_link_input)
        )

        st.markdown("")
        st.markdown("**Folgende Unterordner werden automatisch erstellt:**")
        subfolder_structure = [
            "01_Projektunterlagen",
            "02_Angebote",
            "03_Rechnungen",
            "04_Vergabeunterlagen",
            "05_Kontaktdaten",
            "06_Nachunternehmen",
            "07_Ausgang"
        ]

        cols = st.columns(4)
        for idx, subfolder in enumerate(subfolder_structure):
            with cols[idx % 4]:
                st.markdown(f"✓ {subfolder}")

        st.markdown("")

        # Cloud environment: Create ZIP file
        if in_cloud:
            # Prepare ZIP file
            zip_buffer = io.BytesIO()
            project_name_for_folder = st.session_state.folder_name_input
            safe_main_folder = sanitize_filename(project_name_for_folder)

            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                # Create empty folders without placeholder files
                # Don't use project name prefix - it's already in the ZIP filename
                for subfolder in subfolder_structure:
                    folder_path = f"{subfolder}/"
                    # Create the folder entry in the ZIP (empty directory)
                    zip_info = zipfile.ZipInfo(folder_path)
                    zip_file.writestr(zip_info, "")

                # Always add project link file at root of ZIP
                link_filename = "Projekt_Link.txt"
                if st.session_state.project_link and st.session_state.project_link.strip():
                    link_content = f"Projekt-Link:\n{st.session_state.project_link}\n\nProjektname: {project_name_for_folder}\nErstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                else:
                    link_content = f"Projektname: {project_name_for_folder}\nErstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\nHinweis: Kein Projekt-Link angegeben."
                zip_file.writestr(link_filename, link_content)

            zip_filename = f"{safe_main_folder}.zip"

            col_zip1, col_zip2 = st.columns(2)
            with col_zip1:
                st.download_button(
                    label="📦 Ordnerstruktur als ZIP herunterladen",
                    data=zip_buffer.getvalue(),
                    file_name=zip_filename,
                    mime="application/zip",
                    use_container_width=True,
                    type="primary",
                    key="download_folder_structure"
                )

            # Show what's included and provide separate project link download
            with col_zip2:
                st.markdown("**📋 Im ZIP enthalten:**")
                st.markdown(f"✓ {len(subfolder_structure)} Unterordner")
                st.markdown("✓ Projekt_Link.txt")
                if st.session_state.project_link and st.session_state.project_link.strip():
                    st.success("✓ Mit Projekt-Link")
                else:
                    st.info("ℹ️ Ohne Projekt-Link")

            # Additional option: Download project link file separately
            if st.session_state.project_link and st.session_state.project_link.strip():
                st.markdown("")
                link_file_content = f"Projekt-Link:\n{st.session_state.project_link}\n\nProjektname: {project_name_for_folder}\nErstellt am: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                st.download_button(
                    label="📄 Nur Projekt-Link.txt herunterladen",
                    data=link_file_content,
                    file_name=f"Projekt_Link_{safe_main_folder}.txt",
                    mime="text/plain",
                    use_container_width=False,
                    key="download_project_link_only"
                )

        # Local environment: Create folders directly
        else:
            if st.button("🗂️ Ordnerstruktur erstellen", use_container_width=True, type="primary", key="create_folders"):
                try:
                    project_name_for_folder = st.session_state.folder_name_input
                    safe_main_folder = sanitize_filename(project_name_for_folder)
                    folder_location = st.session_state.folder_location
                    main_folder_path = os.path.join(folder_location, safe_main_folder)

                    if not os.path.exists(folder_location):
                        st.error(f"❌ Der Speicherort existiert nicht: {folder_location}")
                    elif os.path.exists(main_folder_path):
                        st.warning(f"⚠️ Der Ordner existiert bereits: {main_folder_path}")
                    else:
                        # Create main folder
                        os.makedirs(main_folder_path, exist_ok=True)

                        # Create subfolders
                        for subfolder in subfolder_structure:
                            subfolder_path = os.path.join(main_folder_path, subfolder)
                            os.makedirs(subfolder_path, exist_ok=True)

                        # Create project link text file if link is provided
                        if st.session_state.project_link and st.session_state.project_link.strip():
                            link_filename = f"Projekt_Link_{safe_main_folder}.txt"
                            link_filepath = os.path.join(main_folder_path, link_filename)
                            with open(link_filepath, 'w', encoding='utf-8') as f:
                                f.write(f"{st.session_state.project_link}\n")

                        st.success(f"✅ **Erfolgreich erstellt!**")
                        if st.session_state.project_link and st.session_state.project_link.strip():
                            st.success(f"✅ **Projekt-Link gespeichert!**")
                        st.info(f"📂 Hauptordner: `{main_folder_path}`")

                except PermissionError:
                    st.error(f"❌ Keine Berechtigung zum Erstellen von Ordnern in: {folder_location}")
                except Exception as e:
                    st.error(f"❌ Fehler beim Erstellen der Ordnerstruktur: {str(e)}")
    
    
    # Step 3: Export
    st.markdown("---")
    st.subheader("📥 Schritt 3: Angebot exportieren")

    # Project name input for all environments
    st.markdown("")
    export_filename_base = st.text_input(
        "📝 Projektname für Export:",
        value=st.session_state.project_name if st.session_state.project_name else "Bauprojekt",
        help="Dieser Name wird für die exportierten Dateien verwendet",
        key="export_project_name"
    )
    st.session_state.project_name = export_filename_base

    # Check environment for different export methods
    in_cloud = is_cloud_environment()

    if in_cloud:
        st.info("💡 Dateien werden direkt an Ihren Browser zum Download gesendet.")
    else:
        st.markdown("Dateien werden automatisch im angegebenen Speicherort mit dem Projektnamen gespeichert.")
        export_folder = st.session_state.folder_location

    st.markdown("")
    col1, col2 = st.columns(2)
    
    with col1:
        # Excel Export with proper German number formatting
        # Prepare Excel data
        excel_buffer = io.BytesIO()

        # Prepare export dataframe with German-formatted text
        export_df = edited_df[['pos', 'description', 'quantity', 'unit', 'unit_price']].copy()
        # Calculate GP (Menge × EP)
        export_df['total_price'] = edited_df['quantity'] * edited_df['unit_price']

        # Convert numeric columns to German-formatted text strings
        export_df['quantity'] = export_df['quantity'].apply(lambda x: format_german_number(x, 2))
        export_df['unit_price'] = export_df['unit_price'].apply(lambda x: format_german_number(x, 2))
        export_df['total_price'] = export_df['total_price'].apply(lambda x: format_german_number(x, 2))

        # Rename columns to German
        export_df.columns = ['Pos.', 'Leistungsbezeichnung', 'Menge', 'Einheit', 'EP netto (€)', 'GP netto (€)']

        # Write to Excel
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            export_df.to_excel(writer, index=False, sheet_name='Kalkulation')

            # Get the worksheet for styling
            worksheet = writer.sheets['Kalkulation']

            # Clean up values and set as text
            from openpyxl.styles import Alignment
            from openpyxl.cell.cell import TYPE_STRING

            for row in range(2, len(export_df) + 2):
                # Menge (column C/3)
                cell_c = worksheet.cell(row=row, column=3)
                clean_value_c = str(cell_c.value).lstrip("'") if cell_c.value else ""
                cell_c.value = clean_value_c
                cell_c.data_type = TYPE_STRING
                cell_c.alignment = Alignment(horizontal='right')

                # EP netto (column E/5)
                cell_e = worksheet.cell(row=row, column=5)
                clean_value_e = str(cell_e.value).lstrip("'") if cell_e.value else ""
                cell_e.value = clean_value_e
                cell_e.data_type = TYPE_STRING
                cell_e.alignment = Alignment(horizontal='right')

                # GP netto (column F/6)
                cell_f = worksheet.cell(row=row, column=6)
                clean_value_f = str(cell_f.value).lstrip("'") if cell_f.value else ""
                cell_f.value = clean_value_f
                cell_f.data_type = TYPE_STRING
                cell_f.alignment = Alignment(horizontal='right')

            # Adjust column widths
            worksheet.column_dimensions['A'].width = 12
            worksheet.column_dimensions['B'].width = 50
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 10
            worksheet.column_dimensions['E'].width = 18
            worksheet.column_dimensions['F'].width = 18

        # Sanitize filename
        safe_filename = sanitize_filename(export_filename_base)
        excel_filename = f"Kalkulation_{safe_filename}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        # Cloud environment: Use download button
        if in_cloud:
            st.download_button(
                label="💾 Excel herunterladen",
                data=excel_buffer.getvalue(),
                file_name=excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
        # Local environment: Save to file
        else:
            if st.button("💾 Excel speichern", use_container_width=True, type="primary"):
                try:
                    safe_project_folder = sanitize_filename(export_filename_base)
                    project_folder_path = os.path.join(export_folder, safe_project_folder)

                    if not os.path.exists(export_folder):
                        st.error(f"❌ Der Speicherort existiert nicht: {export_folder}")
                    else:
                        os.makedirs(project_folder_path, exist_ok=True)
                        excel_filepath = os.path.join(project_folder_path, excel_filename)

                        with open(excel_filepath, 'wb') as f:
                            f.write(excel_buffer.getvalue())

                        if st.session_state.project_link and st.session_state.project_link.strip():
                            link_filename = f"Projekt_Link_{safe_filename}.txt"
                            link_filepath = os.path.join(project_folder_path, link_filename)
                            with open(link_filepath, 'w', encoding='utf-8') as f:
                                f.write(f"{st.session_state.project_link}\n")

                        st.success(f"✅ **Excel gespeichert!**")
                        if st.session_state.project_link and st.session_state.project_link.strip():
                            st.success(f"✅ **Projekt-Link gespeichert!**")
                        st.info(f"📂 Speicherort: `{excel_filepath}`")

                except PermissionError:
                    st.error(f"❌ Keine Berechtigung zum Schreiben in: {export_folder}")
                except Exception as e:
                    st.error(f"❌ Excel-Fehler: {str(e)}")
    
    with col2:
        # PDF Export
        # Prepare PDF data
        with st.spinner("Erstelle PDF..."):
            try:
                # Prepare dataframe for PDF with calculated total prices
                pdf_df = edited_df.copy()
                pdf_df['total_price'] = pdf_df['quantity'] * pdf_df['unit_price']

                # Generate PDF
                pdf_bytes = generate_offer_pdf(pdf_df, export_filename_base)

                # Sanitize filename
                safe_filename = sanitize_filename(export_filename_base)
                pdf_filename = f"Angebot_{safe_filename}_{datetime.now().strftime('%Y%m%d')}.pdf"

                # Cloud environment: Use download button
                if in_cloud:
                    st.download_button(
                        label="📄 PDF herunterladen",
                        data=pdf_bytes,
                        file_name=pdf_filename,
                        mime="application/pdf",
                        use_container_width=True,
                        type="primary"
                    )
                # Local environment: Save to file with button
                else:
                    # Store pdf_bytes in session state for the button
                    if 'pdf_data' not in st.session_state:
                        st.session_state.pdf_data = pdf_bytes
                        st.session_state.pdf_filename = pdf_filename

            except Exception as e:
                st.error(f"❌ PDF-Fehler beim Erstellen: {str(e)}")

        # Local save button (outside spinner)
        if not in_cloud:
            if st.button("📄 PDF speichern", use_container_width=True, type="primary", key="pdf_save_btn"):
                try:
                    safe_project_folder = sanitize_filename(export_filename_base)
                    project_folder_path = os.path.join(export_folder, safe_project_folder)

                    if not os.path.exists(export_folder):
                        st.error(f"❌ Der Speicherort existiert nicht: {export_folder}")
                    else:
                        os.makedirs(project_folder_path, exist_ok=True)
                        pdf_filepath = os.path.join(project_folder_path, st.session_state.pdf_filename)

                        with open(pdf_filepath, 'wb') as f:
                            f.write(st.session_state.pdf_data)

                        if st.session_state.project_link and st.session_state.project_link.strip():
                            link_filename = f"Projekt_Link_{safe_filename}.txt"
                            link_filepath = os.path.join(project_folder_path, link_filename)
                            with open(link_filepath, 'w', encoding='utf-8') as f:
                                f.write(f"{st.session_state.project_link}\n")

                        st.success(f"✅ **PDF gespeichert!**")
                        if st.session_state.project_link and st.session_state.project_link.strip():
                            st.success(f"✅ **Projekt-Link gespeichert!**")
                        st.info(f"📂 Speicherort: `{pdf_filepath}`")

                except PermissionError:
                    st.error(f"❌ Keine Berechtigung zum Schreiben in: {export_folder}")
                except Exception as e:
                    st.error(f"❌ PDF-Fehler: {str(e)}")

else:
    st.markdown("""
    <div style='text-align: center; padding: 40px; background-color: #f9fafb; border-radius: 10px; margin: 20px 0;'>
        <h3 style='color: #6b7280;'>📋 Keine Daten vorhanden</h3>
        <p style='color: #9ca3af; font-size: 1.1em;'>
            Bitte laden Sie zuerst ein Dokument hoch, um mit der Bearbeitung zu beginnen.<br>
            Unterstützte Formate: GAEB, PDF, Excel, Word
        </p>
    </div>
    """, unsafe_allow_html=True)

# Footer with enhanced styling
st.markdown("---")
st.markdown(f"""
<div style='text-align: center; padding: 20px; color: #666;'>
    <p style='margin: 5px 0;'><strong>{COMPANY_NAME}</strong></p>
    <p style='margin: 5px 0; font-size: 0.85em; color: #999;'>© {datetime.now().year} Alle Rechte vorbehalten</p>
</div>
""", unsafe_allow_html=True)
