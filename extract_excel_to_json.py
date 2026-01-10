"""
Excel to JSON Extraction Tool
Extracts position data from Excel files and converts to JSON format
"""

import openpyxl
import json
from pathlib import Path


def extract_positions_from_excel(excel_file_path, output_json_path=None):
    """
    Extracts position data from Excel file where column A contains "Position"
    
    Args:
        excel_file_path: Path to the Excel file
        output_json_path: Optional path to save JSON output (if None, just returns data)
    
    Returns:
        List of dictionaries containing position data
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    
    positions = []
    
    # Iterate through all rows in column A
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet[f'A{row}'].value
        
        # Check if cell contains "Position"
        if cell_value and "Position" in str(cell_value):
            # Extract data from the same row
            position_data = {
                "Typ": "Position",
                "Ordnungszahl": sheet[f'B{row}'].value,
                "Kurztext": sheet[f'C{row}'].value,
                "Langtext": sheet[f'D{row}'].value,
                "Menge": sheet[f'E{row}'].value,
                "Einheit": sheet[f'F{row}'].value
            }
            
            # Clean up the data (remove None values, convert types)
            cleaned_data = {}
            for key, value in position_data.items():
                if value is not None:
                    # Convert to appropriate type
                    if key == "Menge":
                        try:
                            cleaned_data[key] = float(value) if '.' in str(value) else int(value)
                        except (ValueError, TypeError):
                            cleaned_data[key] = value
                    else:
                        cleaned_data[key] = str(value).strip() if isinstance(value, str) else value
                else:
                    cleaned_data[key] = value
            
            positions.append(cleaned_data)
            print(f"Found Position at row {row}: {cleaned_data.get('Kurztext', 'N/A')}")
    
    # Save to JSON file if path provided
    if output_json_path:
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(positions, f, ensure_ascii=False, indent=2)
        print(f"\nData saved to {output_json_path}")
    
    return positions


def analyze_excel_structure(excel_file_path):
    """
    Analyzes the Excel file structure to understand the layout
    Prints the first few rows to help understand the data structure
    """
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    
    print(f"Analyzing Excel file: {excel_file_path}")
    print(f"Sheet name: {sheet.title}")
    print(f"Max row: {sheet.max_row}, Max column: {sheet.max_column}")
    print("\nFirst 20 rows (columns A-F):")
    print("-" * 100)
    
    for row in range(1, min(21, sheet.max_row + 1)):
        row_data = []
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell_value = sheet[f'{col}{row}'].value
            row_data.append(str(cell_value)[:30] if cell_value else "")
        print(f"Row {row:2d}: {' | '.join(row_data)}")
    
    print("-" * 100)


def main():
    # Define the Excel file path
    excel_file = Path("Data/LV 08 Trockenbau (STEP)_20251205_GAEB (1) (1).xlsx")
    
    if not excel_file.exists():
        print(f"Error: Excel file not found at {excel_file}")
        return
    
    print("=" * 100)
    print("EXCEL STRUCTURE ANALYSIS")
    print("=" * 100)
    analyze_excel_structure(excel_file)
    
    print("\n" + "=" * 100)
    print("EXTRACTING POSITION DATA")
    print("=" * 100)
    
    # Extract positions (no file output)
    positions = extract_positions_from_excel(excel_file)
    
    print(f"\nTotal positions found: {len(positions)}")
    
    # Display the results
    if positions:
        print("\n" + "=" * 100)
        print("EXTRACTED DATA (JSON format)")
        print("=" * 100)
        print(json.dumps(positions, ensure_ascii=False, indent=2))
    else:
        print("\nNo positions found in the Excel file.")

    print(f"\nTotal positions found: {len(positions)}")


if __name__ == "__main__":
    main()
